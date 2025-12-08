# Takes input as array and adds it as a new column to an existing excel file
def add_quantity(z):

    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Adding QUANTITY column
    ws.cell(row=1, column=3, value="QUANTITY")

    row_1 = 2
    for i in range(len(z)):
        ws.cell(row=row_1, column=3, value=z[i])
        row_1 += 1

    wb.save(EXCEL_FILE)

# Takes input as array and adds it as a new column to an existing excel file
def add_rate(z):

    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Adding RATE column
    ws.cell(row=1, column=4, value="RATE")

    row_1 = 2
    for i in range(len(z)):
        ws.cell(row=row_1, column=4, value=z[i])
        row_1 += 1

    wb.save(EXCEL_FILE)