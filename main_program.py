import cv2 as cv
import numpy as np
import os

# detect red corners
def get_red_corners(img):

    hsv = cv.cvtColor(img, cv.COLOR_BGR2HSV) 

    # Red color ranges
    lower_red1 = np.array([0, 120, 70])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 120, 70])
    upper_red2 = np.array([180, 255, 255])

    mask = cv.inRange(hsv, lower_red1, upper_red1) + cv.inRange(hsv, lower_red2, upper_red2)

    contours, _ = cv.findContours(mask, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)
    corners = []
    for cnt in contours:
        M = cv.moments(cnt)
        if M["m00"] != 0:
            cx = int(M["m10"]/M["m00"])
            cy = int(M["m01"]/M["m00"])
            corners.append([cx, cy])

    if len(corners) != 4:
        raise ValueError(f"Expected 4 red corners, found {len(corners)}")
    
    return np.array(corners, dtype=np.float32)

# detect four corners in the hard copy
def get_paper_corners(img):

    import cv2 as cv
    import numpy as np
    import os

    gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
    blur = cv.GaussianBlur(gray, (5,5), 0)
    edges = cv.Canny(blur, 50, 150)

    contours, _ = cv.findContours(edges, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)
    if not contours:
        raise ValueError("No contours found in hard copy")
    cnt = max(contours, key=cv.contourArea)
    peri = cv.arcLength(cnt, True)
    approx = cv.approxPolyDP(cnt, 0.02 * peri, True)

    if len(approx) != 4:
        raise ValueError(f"Expected 4 corners in hard copy, found {len(approx)}")
    
    return np.array([p[0] for p in approx], dtype=np.float32)

# top left, bottom right, top right and bottom left
def order_points(pts):

    import cv2 as cv
    import numpy as np
    import os

    rect = np.zeros((4,2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]  # TL
    rect[2] = pts[np.argmax(s)]  # BR
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]  # TR
    rect[3] = pts[np.argmax(diff)]  # BL
    return rect

# to extract the table from the hard copy 
def crop_image_edges_cv(img, top=0, bottom=0, left=0, right=0):

    import cv2 as cv
    import numpy as np
    import os

    h, w = img.shape[:2]

    # Compute valid crop boundaries
    y1 = top
    y2 = h - bottom
    x1 = left
    x2 = w - right

    # Safety clamp to prevent invalid cropping
    y1 = max(0, min(h, y1))
    y2 = max(0, min(h, y2))
    x1 = max(0, min(w, x1))
    x2 = max(0, min(w, x2))

    return img[y1:y2, x1:x2]

# to extract cells from the table

def save_table_cells_to_folder(table_img, col_widths, col_names,
                               folder, row_height, thickness):
    
    import cv2 as cv
    import numpy as np
    import os

    """
    table_img  : input table image (numpy array)
    col_widths : list of column widths in pixels
    col_names  : list of column names
    folder     : folder to save images
    row_height : fixed height of each table row (default 57px)
    thickness  : border pixels to remove from all 4 sides (default 5px)
    """
    os.makedirs(folder, exist_ok=True)

    height, width = table_img.shape[:2]
    n_rows = height // row_height  # compute rows using known row height

    for r in range(n_rows):
        # Row cropping
        y1 = r * row_height
        y2 = y1 + row_height
        row_img = table_img[y1:y2, :]

        x_start = 0
        for c, w in enumerate(col_widths):
            x1 = x_start
            x2 = x_start + w
            cell_img = row_img[:, x1:x2]

            # ---- Remove border thickness ----
            t = thickness
            h, w_cell = cell_img.shape[:2]

            cell_img = cell_img[
                t : h - t,
                t : w_cell - t
            ]

            # Save
            filename = os.path.join(folder, f"{col_names[c]}_{r+1}.jpg")
            cv.imwrite(filename, cell_img)

            x_start += w

# to stitch the cells of the particulars column into one
def extract_column(img, filename, distance_from_left, column_width):

    import cv2 as cv
    import numpy as np
    import os

    h, w = img.shape[:2]

    x1 = distance_from_left
    x2 = distance_from_left + column_width

    # Ensure safe bounds
    x1 = max(0, min(w, x1))
    x2 = max(0, min(w, x2))

    cv.imwrite(filename, img[:, x1:x2])


def extract_form_data(input_dir, model_path, num_items=16):

    import cv2
    import numpy as np
    import tensorflow as tf
    from tensorflow.keras.models import load_model
    import os

    """
    Main pipeline function to extract digit data from cropped cell images.
    
    ARGS:
        input_dir (str): Path to the folder containing cropped cell images (e.g., 'slno_1.jpg').
        model_path (str): File path to the trained .keras model.
        num_items (int): Number of rows/items to process (default 16).

    RETURNS:
        tuple: (sl_no_array, rate_array, qty_array) 
               Three numpy arrays containing the extracted integers.
    """
    
    # Load the trained model once at the start
    try:
        model = load_model(model_path)
        print(f"Model loaded successfully from: {model_path}")
    except Exception as e:
        print(f"Error loading model: {e}")
        # Return empty arrays in case of model failure
        return np.array([]), np.array([]), np.array([])

    # --- Helper Functions (Internal Logic) ---

    def clean_cell_borders(img, margin=4):
        h, w = img.shape[:2]
        cv2.rectangle(img, (0, 0), (w, h), (255, 255, 255), thickness=margin*2)
        return img

    def preprocess_digit_for_mnist(digit_roi):
        if np.mean(digit_roi) > 127:
            digit_roi = cv2.bitwise_not(digit_roi)
        
        coords = cv2.findNonZero(digit_roi)
        if coords is None:
            return np.zeros((28, 28, 1))
            
        x, y, w, h = cv2.boundingRect(coords)
        roi = digit_roi[y:y+h, x:x+w]
        
        rows, cols = roi.shape
        if rows > cols:
            factor = 20.0 / rows
            rows = 20
            cols = int(round(cols * factor))
        else:
            factor = 20.0 / cols
            cols = 20
            rows = int(round(rows * factor))
            
        roi = cv2.resize(roi, (cols, rows), interpolation=cv2.INTER_AREA)
        
        colsPadding = (int(np.ceil((28 - cols) / 2.0)), int(np.floor((28 - cols) / 2.0)))
        rowsPadding = (int(np.ceil((28 - rows) / 2.0)), int(np.floor((28 - rows) / 2.0)))
        padded = np.pad(roi, (rowsPadding, colsPadding), 'constant')
        
        _, padded = cv2.threshold(padded, 30, 255, cv2.THRESH_BINARY)
        
        padded = padded.astype('float32') / 255.0
        padded = np.expand_dims(padded, axis=-1)
        return padded

    def process_single_cell(image_path, expected_val=None):
        if not os.path.exists(image_path): return 0
            
        img = cv2.imread(image_path)
        if img is None: return 0
        
        img = clean_cell_borders(img, margin=3)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
        
        cnts, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if not cnts: return 0
            
        boundingBoxes = [cv2.boundingRect(c) for c in cnts]
        (cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes), key=lambda b: b[1][0]))
        
        batch_digits = []
        
        for i, c in enumerate(cnts):
            x, y, w, h = cv2.boundingRect(c)
            if h < 12 or w < 3: continue
                
            if w > 1.3 * h:
                half_w = w // 2
                batch_digits.append(preprocess_digit_for_mnist(thresh[y:y+h, x:x+half_w]))
                batch_digits.append(preprocess_digit_for_mnist(thresh[y:y+h, x+half_w:x+w]))
            else:
                batch_digits.append(preprocess_digit_for_mnist(thresh[y:y+h, x:x+w]))
            
        if len(batch_digits) > 0:
            predictions = model.predict(np.array(batch_digits), verbose=0)
            
            if expected_val is not None:
                expected_str = str(expected_val)
                if len(predictions) == len(expected_str):
                    for k in range(len(predictions)):
                        target_digit = int(expected_str[k])
                        predictions[k][target_digit] += 2.0
    
            predicted_digits = ""
            for p in predictions:
                predicted_digits += str(np.argmax(p))
                    
            return int(predicted_digits) if predicted_digits else 0
        
        return 0

    # --- Main Processing Loop ---
    sl_no_list = []
    rate_list = []
    qty_list = []
    
    print(f"Processing {num_items} items from: {input_dir}")
    
    for i in range(1, num_items + 1):
        try:
            # Dynamically join the input_dir with standard filenames
            path_sl = os.path.join(input_dir, f"slno_{i}.jpg")
            path_rate = os.path.join(input_dir, f"rate_{i}.jpg")
            path_qty = os.path.join(input_dir, f"quantity_{i}.jpg")
            
            val_sl = process_single_cell(path_sl, expected_val=i)
            val_rate = process_single_cell(path_rate, expected_val=None)
            val_qty = process_single_cell(path_qty, expected_val=None)
            
            sl_no_list.append(val_sl)
            rate_list.append(val_rate)
            qty_list.append(val_qty)
            
        except Exception as e:
            print(f"Error processing row {i}: {e}")
            sl_no_list.append(0)
            rate_list.append(0)
            qty_list.append(0)

    return sl_no_list, rate_list, qty_list

def manage_grocery_list(action, image_path=None, item_name=None):

    import os
    import difflib
    import pytesseract
    from openpyxl import Workbook, load_workbook
    import cv2
    import numpy as np
    pytesseract.pytesseract.tesseract_cmd = (r"C:\Program Files\Tesseract-OCR\tesseract.exe")
    EXCEL_FILE = "initial.xlsx"

    VALID_ITEMS = [
        "Rice","Wheat","Flour","Sugar","Salt","Milk","Eggs","Butter","Ghee","Cheese",
        "Yogurt","Bread","Oats","Cornflakes","Pasta","Noodles","Tea","Coffee","Green Tea",
        "Honey","Jam","Peanut Butter","Cooking Oil","Olive Oil","Sunflower Oil","Coconut Oil",
        "Mustard Oil","Turmeric Powder","Chili Powder","Coriander Powder","Cumin Seeds",
        "Black Pepper","Cardamom","Cloves","Cinnamon","Bay Leaf","Garam Masala","Sambar Powder",
        "Rasam Powder","Lentils","Toor Dal","Moong Dal","Chana Dal","Urad Dal","Rajma",
        "Chickpeas","Green Peas","Dry Fruits Mix","Almonds","Cashews","Raisins","Walnuts",
        "Pistachios","Biscuits","Cookies","Chips","Popcorn","Chocolate","Candy","Tomato Ketchup",
        "Soy Sauce","Vinegar","Mayonnaise","Pickles","Onions","Potatoes","Tomatoes","Carrots",
        "Beans","Cabbage","Cauliflower","Spinach","Ginger","Garlic","Green Chillies","Apples",
        "Bananas","Oranges","Grapes","Mangoes","Papaya","Pomegranate","Watermelon","Lemon",
        "Cucumber","Soap","Shampoo","Toothpaste","Toothbrush","Detergent","Dishwashing Liquid",
        "Handwash","Sanitizer","Aluminum Foil","Tissue Paper","Garbage Bags","Matches",
        "Candles","Salted Butter","Brown Sugar","Baking Soda","Baking Powder","Yeast","Chips","Fruits"
    ]

    def smart_autocorrect(word):
        clean_word = word.strip()
        for item in VALID_ITEMS:
            if item.lower() == clean_word.lower():
                return item
        matches = difflib.get_close_matches(clean_word, VALID_ITEMS, n=1, cutoff=0.6)
        return matches[0] if matches else clean_word

    try:
        if action == "extract_from_image":
            if not image_path or not os.path.exists(image_path):
                return f"Error: Could not find image at: {image_path}"

            img = cv2.imread(image_path)
            if img is None:
                return "Error: Cv2 could not read image."

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            thresh = cv2.bitwise_not(gray)
            bw = cv2.adaptiveThreshold(thresh, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                    cv2.THRESH_BINARY, 15, -2)

            cols = bw.shape[1]
            horizontal_size = max(1, cols // 30)
            horizontalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (horizontal_size, 1))
            horizontal = cv2.erode(bw, horizontalStructure)
            horizontal = cv2.dilate(horizontal, horizontalStructure)

            contours, _ = cv2.findContours(horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

            lines_y = []
            for cnt in contours:
                x, y, w, h = cv2.boundingRect(cnt)
                if w > cols / 3:
                    lines_y.append(y)

            lines_y.sort()
            lines_y = [0] + lines_y + [img.shape[0]]

            extracted_words = []

            for i in range(len(lines_y) - 1):
                y1 = lines_y[i]
                y2 = lines_y[i + 1]

                roi = img[y1:y2, 0:cols]
                if roi.shape[0] < 15:
                    continue

                # Convert to grayscale and binarize for better OCR
                gray_roi = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
                _, roi_bin = cv2.threshold(gray_roi, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

                text = pytesseract.image_to_string(roi_bin, config='--psm 6').strip()
                print(f"OCR Line {i+1}: '{text}'")  # For debugging

                if text:
                    extracted_words.append(text)

            if os.path.exists(EXCEL_FILE):
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active

            ws.cell(row=1, column=1, value="PARTICULARS")

            start_row = 2
            items_added = 0

            for word in extracted_words:
                clean_txt = "".join([c for c in word if c.isalpha() or c.isspace()])
                if clean_txt.strip():
                    final_word = smart_autocorrect(clean_txt).title()
                    ws.cell(row=start_row, column=1, value=final_word)
                    start_row += 1
                    items_added += 1

            wb.save(EXCEL_FILE)
            return f"Success: Processed image via improved OCR. Added {items_added} items."

        elif action == "delete_item":
            if not item_name: return "Error: Item name missing."
            if not os.path.exists(EXCEL_FILE): return "Error: List file not found."

            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            deleted_count = 0
            
            for row in range(ws.max_row, 1, -1):
                val = ws.cell(row=row, column=1).value
                if val and str(val).lower() == item_name.lower():
                    ws.delete_rows(row)
                    deleted_count += 1
            
            wb.save(EXCEL_FILE)
            return f"Success: Deleted {deleted_count} occurrences of '{item_name}'."

        elif action == "delete_all":
            if os.path.exists(EXCEL_FILE):
                os.remove(EXCEL_FILE)
                return "Success: Full list deleted."
            return "Info: List was already empty."

        else:
            return "Error: Unknown action."

    except PermissionError:
        return "Error: Close the Excel file and try again."
    except Exception as e:
        return f"System Error: {str(e)}"
    
# Takes input as array and adds it as a new column to an existing excel file
def add_quantity(z):

    from openpyxl import load_workbook

    wb = load_workbook("initial.xlsx")
    ws = wb.active

    # Adding QUANTITY column
    ws.cell(row=1, column=2, value="QUANTITY")

    row_1 = 2
    for i in range(len(z)):
        ws.cell(row=row_1, column=2, value=z[i])
        row_1 += 1

    wb.save("initial.xlsx")

# Takes input as array and adds it as a new column to an existing excel file
def add_rate(z):

    from openpyxl import load_workbook

    wb = load_workbook("initial.xlsx")
    ws = wb.active

    # Adding RATE column
    ws.cell(row=1, column=3, value="RATE")

    row_1 = 2
    for i in range(len(z)):
        ws.cell(row=row_1, column=3, value=z[i])
        row_1 += 1

    wb.save("initial.xlsx")

# Converts an excel file to csv format
def excel_to_csv(y):
    import pandas as pd

    df = pd.read_excel(y)

    df.to_csv("output.csv", index=False)

# Converts a csv file to an excel file with additional computed columns and formatting
def csv_to_excel(x):

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(x)
    df.columns = df.columns.str.strip()

    # Aggregate QUANTITY by PARTICULARS and keep the first RATE
    df = df.groupby(["PARTICULARS"], as_index=False).agg({
        "QUANTITY": "sum",
        "RATE": "first"
    })

    # Add SL NO. and additional columns
    df.insert(0, "SL NO.", range(1, len(df) + 1))

    df["EXISTING QUANTITY"] = ""
    df["AMOUNT"] = ""
    df["REMAINING QUANTITY"] = ""
    df["STATUS"] = ""

    # Reorder columns
    df = df[["SL NO.", "PARTICULARS", "QUANTITY", "RATE", "AMOUNT","EXISTING QUANTITY", "REMAINING QUANTITY", "STATUS"]]

    output_file = "items.xlsx"
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    # Adding formulas for AMOUNT, REMAINING QUANTITY, and STATUS columns
    for row in range(2, ws.max_row + 1):  
        ws[f"E{row}"].value = f"=C{row}*D{row}"
        ws[f"G{row}"].value = f"=F{row}-C{row}"
        ws[f"H{row}"].value = f'=IF(G{row}<=10,"LOW STOCK","SUFFICIENT STOCK")' 

    last = ws.max_row + 1
    ws[f"D{last}"] = "TOTAL"
    ws[f"E{last}"] = f"=SUM(E2:E{last-1})"

    # Adds color formatting
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # yellow
    total_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")   # light green

    for cell in ws[1]:
        cell.fill = header_fill

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.fill = total_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter

        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_file)

# -----------------------------
# MAIN PROGRAM
# -----------------------------

soft = cv.imread("soft_copy_template.jpg")
hard = cv.imread("hard_copy_photo.jpg")

if soft is None or hard is None:
    raise ValueError("Check image paths! Images not found.")

# Get corners
soft_corners = order_points(get_red_corners(soft))
hard_corners = order_points(get_paper_corners(hard))

# Compute homography and warp
H, _ = cv.findHomography(hard_corners, soft_corners)
aligned = cv.warpPerspective(hard, H, (soft.shape[1], soft.shape[0]))
final = crop_image_edges_cv(aligned, top=90, bottom=45, left=34, right=28)

# Load table image
table = final

# Row heights (in pixels) — specify exact heights for each row
n_rows = 17

# Column widths (in pixels)
col_widths = [42, 365, 124, 124]

# Column names
col_names = ["slno", "particulars", "rate", "quantity"]

# Save the particulars column
extract_column(final, "cells/particulars.jpg", 47, 365)

# Save all cells
folder = "cells"
save_table_cells_to_folder(table, col_widths, col_names, folder, 55, 5)

# Extract data using the trained model
sl, rate, qty = extract_form_data("cells", "model.keras", num_items=16)

manage_grocery_list("extract_from_image", image_path="cells/particulars.jpg")
add_rate(rate)
add_quantity(qty)
excel_to_csv("initial.xlsx")
csv_to_excel("output.csv")