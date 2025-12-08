# Converts a csv file to an excel file with additional computed columns and formatting
def csv_to_excel(x):

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(x)
    df.columns = df.columns.str.strip()

    # Aggregate QUANTITY by PARTICULARS and keep the first RATE
    df = df.groupby(["PARTICULARS"], as_index=False).agg({
        "QUANTITY": "sum",
        "RATE": "first"
    })

    # Add SL NO. and additional columns
    df.insert(0, "SL NO.", range(1, len(df) + 1))

    df["EXISTING QUANTITY"] = ""
    df["AMOUNT"] = ""
    df["REMAINING QUANTITY"] = ""
    df["STATUS"] = ""

    # Reorder columns
    df = df[["SL NO.", "PARTICULARS", "QUANTITY", "RATE", "AMOUNT","EXISTING QUANTITY", "REMAINING QUANTITY", "STATUS"]]

    output_file = "items.xlsx"
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    # Adding formulas for AMOUNT, REMAINING QUANTITY, and STATUS columns
    for row in range(2, ws.max_row + 1):  
        ws[f"E{row}"].value = f"=C{row}*D{row}"
        ws[f"G{row}"].value = f"=F{row}-C{row}"
        ws[f"H{row}"].value = f'=IF(G{row}<=10,"LOW STOCK","SUFFICIENT STOCK")' 

    last = ws.max_row + 1
    ws[f"D{last}"] = "TOTAL"
    ws[f"E{last}"] = f"=SUM(E2:E{last-1})"

    # Adds color formatting
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # yellow
    total_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")   # light green

    for cell in ws[1]:
        cell.fill = header_fill

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.fill = total_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter

        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_file)