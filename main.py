import os
import difflib
import pytesseract
from openpyxl import Workbook, load_workbook
import cv2
import numpy as np
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
global EXCEL_FILE
EXCEL_FILE = r'C:\Users\aksps\OneDrive\Python\Jackfruit\PARTICULARS.xlsx'
DEFAULT_IMAGE_PATH = r'C:\Users\aksps\OneDrive\Python\Jackfruit\Images\tds.jpg'

def manage_grocery_list(action, image_path=None, item_name=None):

    VALID_ITEMS = [
        "Rice","Wheat","Flour","Sugar","Salt","Milk","Eggs","Butter","Ghee","Cheese",
        "Yogurt","Bread","Oats","Cornflakes","Pasta","Noodles","Tea","Coffee","Green Tea",
        "Honey","Jam","Peanut Butter","Cooking Oil","Olive Oil","Sunflower Oil","Coconut Oil",
        "Mustard Oil","Turmeric Powder","Chili Powder","Coriander Powder","Cumin Seeds",
        "Black Pepper","Cardamom","Cloves","Cinnamon","Bay Leaf","Garam Masala","Sambar Powder",
        "Rasam Powder","Lentils","Toor Dal","Moong Dal","Chana Dal","Urad Dal","Rajma",
        "Chickpeas","Green Peas","Dry Fruits Mix","Almonds","Cashews","Raisins","Walnuts",
        "Pistachios","Biscuits","Cookies","Chips","Popcorn","Chocolate","Candy","Tomato Ketchup",
        "Soy Sauce","Vinegar","Mayonnaise","Pickles","Onions","Potatoes","Tomatoes","Carrots",
        "Beans","Cabbage","Cauliflower","Spinach","Ginger","Garlic","Green Chillies","Apples",
        "Bananas","Oranges","Grapes","Mangoes","Papaya","Pomegranate","Watermelon","Lemon",
        "Cucumber","Soap","Shampoo","Toothpaste","Toothbrush","Detergent","Dishwashing Liquid",
        "Handwash","Sanitizer","Aluminum Foil","Tissue Paper","Garbage Bags","Matches",
        "Candles","Salted Butter","Brown Sugar","Baking Soda","Baking Powder","Yeast","Chips","Fruits"
    ]

    def smart_autocorrect(word):
        clean_word = word.strip()
        for item in VALID_ITEMS:
            if item.lower() == clean_word.lower():
                return item
        matches = difflib.get_close_matches(clean_word, VALID_ITEMS, n=1, cutoff=0.6)
        return matches[0] if matches else clean_word

    try:
        if action == "extract_from_image":
            if not image_path or not os.path.exists(image_path):
                return f"Error: Could not find image at: {image_path}"

            img = cv2.imread(image_path)
            if img is None: return "Error: Cv2 could not read image."

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            thresh = cv2.bitwise_not(gray)
            bw = cv2.adaptiveThreshold(thresh, 255, cv2.ADAPTIVE_THRESH_MEAN_C, 
                                       cv2.THRESH_BINARY, 15, -2)

            cols = bw.shape[1]
            horizontal_size = cols // 30
            horizontalStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (horizontal_size, 1))
            horizontal = cv2.erode(bw, horizontalStructure)
            horizontal = cv2.dilate(horizontal, horizontalStructure)

            contours, _ = cv2.findContours(horizontal, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

            lines_y = []
            for cnt in contours:
                x, y, w, h = cv2.boundingRect(cnt)
                if w > cols / 3:
                    lines_y.append(y)
            
            lines_y.sort()

            lines_y = [0] + lines_y + [img.shape[0]]
            
            extracted_words = []

            for i in range(len(lines_y) - 1):
                y1 = lines_y[i]
                y2 = lines_y[i+1]

                roi = img[y1+5:y2-5, 0:cols]
                if roi.shape[0] < 15: continue
                text = pytesseract.image_to_string(roi, config='--psm 7').strip()
                
                if text:
                    extracted_words.append(text)

            if os.path.exists(EXCEL_FILE):
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                
            start_row = ws.max_row
            items_added = 0
            
            for word in extracted_words:
                clean_txt = "".join([c for c in word if c.isalpha() or c.isspace()])
                if clean_txt.strip():
                    final_word = smart_autocorrect(clean_txt).title()
                    ws.cell(row=start_row, column=2, value=final_word)
                    start_row += 1
                    items_added += 1

            wb.save(EXCEL_FILE)
            return f"Success: Processed image via splitting. Added {items_added} items."

        elif action == "delete_item":
            if not item_name: return "Error: Item name missing."
            if not os.path.exists(EXCEL_FILE): return "Error: List file not found."

            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            deleted_count = 0
            
            for row in range(ws.max_row, 1, -1):
                val = ws.cell(row=row, column=2).value
                if val and str(val).lower() == item_name.lower():
                    ws.delete_rows(row)
                    deleted_count += 1
            
            wb.save(EXCEL_FILE)
            return f"Success: Deleted {deleted_count} occurrences of '{item_name}'."

        elif action == "delete_all":
            if os.path.exists(EXCEL_FILE):
                os.remove(EXCEL_FILE)
                return "Success: Full list deleted."
            return "Info: List was already empty."

        else:
            return "Error: Unknown action."

    except PermissionError:
        return "Error: Close the Excel file and try again."
    except Exception as e:
        return f"System Error: {str(e)}"

if __name__ == "__main__":

    print("\n" + "="*40)
    print(f"PROCESSING DEFAULT IMAGE")
    print(f" Source: {os.path.basename(DEFAULT_IMAGE_PATH)}")
    print("="*40)

    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)

    result = manage_grocery_list("extract_from_image", image_path=DEFAULT_IMAGE_PATH)
    print(result)

    while True:
        print("\n" + "-"*30)
        print(" MENU")
        print("-"*30)
        print("1.Delete Specific Item")
        print("2.Delete Entire List")
        print("3.Exit")
        
        choice = input("\nEnter choice (1-3): ").strip()

        if choice == '1':
            name = input("Enter the item name to delete: ").strip()
            print("\n" + manage_grocery_list("delete_item", item_name=name))

        elif choice == '2':
            print("\n" + manage_grocery_list("delete_all"))
            
        elif choice == '3':
            print("Exiting")
            break
        
        else:
            print("\nInvalid choice, please try again.")